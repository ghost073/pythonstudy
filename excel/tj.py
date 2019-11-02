# -*- utf-8 -*-
# python3.6
# 
# 统计成绩
# 
# 参考：
# Python - openpyxl 读写操作Excel
# https://www.cnblogs.com/BlueSkyyj/p/7571787.html

from openpyxl import load_workbook,Workbook
import os,time
import tjui
import wx

class myui(tjui.MyFrame1):

	# 背景图
	def OnEraseBack( self, event ):
		dc = event.GetDC()
		if not dc:
			dc  = wx.ClientDC(self)
			rect = self.GetUpdateRegion().GetBox()
			dc.SetClippingRect(rect)
		dc.Clear()
		bmp = wx.Bitmap("background.jpg")
		dc.DrawBitmap(bmp, 0, 0)

	# 打开文件
	def openfile( self, event ):
		wildcard = '*.xlsx'
		dialog = wx.FileDialog(None, 'select', os.getcwd(), '', wildcard, wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
			self.infilename.SetValue(dialog.GetPath())
			dialog.Destroy

	# 生成excel操作
	def toexcel( self, event ):
		filename = self.infilename.GetValue()
		# 读取表格信息
		excel_res = self.readExcel(filename)
		# 格式化表格信息
		table_list = self.formatTjTable(excel_res[0], excel_res[1])

		ws_title = [
			{'value':'科目', 'field':'kemu'},
			{'value':'在籍人数', 'field':'people'},
			{'value':'参考人数', 'field':'people'},
			{'value':'总分', 'field':'zongfenscore'},
			{'value':'平均分', 'field':'avgscore'},
			{'value':'60分以上(及格率)', 'field':'jigelv'},
			{'value':'85分以上(优秀率)', 'field':'youxiulv'},
			{'value':'40分以下(低分率)', 'field':'difenlv'},
			{'value':'100及以上', 'field':'fen100'},
			{'value':'99~90', 'field':'fen90_99'},
			{'value':'89~85', 'field':'fen89_85'},
			{'value':'84~80', 'field':'fen84_80'},
			{'value':'79~70', 'field':'fen79_70'},
			{'value':'69~60', 'field':'fen69_60'},
			{'value':'59~50', 'field':'fen59_50'},
			{'value':'49~40', 'field':'fen49_40'},
			{'value':'39~30', 'field':'fen39_30'},
			{'value':'29~20', 'field':'fen29_20'},
			{'value':'19~10', 'field':'fen19_10'},
			{'value':'10分以下', 'field':'fen10xia'},
			{'value':'最低分数', 'field':'minscore'},
		]

		# 创建sheet
		outfilename_val = self.createSheet()

		# 写入表格数据
		self.writeDataExcel(excel_res[0], excel_res[1])
		# 写入表格统计
		self.writeTjExcel(ws_title,table_list)
		# 保存
		self.saveExcel(outfilename_val)

		self.outfilename.SetValue(outfilename_val)
			
	# 创建sheet
	def createSheet(self):
		# 获取文件路径
		nameTime = time.strftime('%Y-%m-%d_%H-%M-%S')
		excelName = 'Excel' + nameTime + '.xlsx'
		ExcelFullName = os.path.join(os.getcwd(), excelName)
		
		self.wb = Workbook()
		self.ws1 = self.wb.create_sheet(title="成绩", index=0)
		self.ws2 = self.wb.create_sheet(title="统计", index=1)

		return ExcelFullName

	# 保存sheet
	def saveExcel(self, fullname):
		self.wb.save(filename=fullname)
		return True

	# 写数据sheet
	def writeDataExcel(self, titlelist, datalist):
		ws = self.ws1

		# 写入标题
		for k,v in enumerate(titlelist):
			c = k + 1
			ws.cell(row=1, column=c).value = v

		# 组装excel 要append的数据
		dataformat = []
		# 列的长度
		max_col = len(datalist[titlelist[0]]['values'])

		zongchengji_title = '总成绩'
		# 追加总成绩列,excel,title,datalist都需要修改
		ws.cell(row=1, column=len(titlelist)+1).value = zongchengji_title
		titlelist.append(zongchengji_title)
		datalist.update({zongchengji_title:{'values':datalist[titlelist[0]]['zongchengji']}})

		for i in range(max_col):
			data_tmp = []
			for titlev in titlelist:
				data_tmp.append(datalist[titlev]['values'][i])
			dataformat.append(data_tmp)

		# 追加excel表数据
		for datav in dataformat:
			ws.append(datav)
		return True	

	# 写统计sheet
	def writeTjExcel(self, titlelist, datalist):

		ws = self.ws2

		for k,v in enumerate(titlelist):
			c = k + 1
			ws.cell(row=1, column=c).value = v['value']

		# 组装excel 要append的数据
		dataformat = []
		for datav in datalist:
			data_tmp = []
			for titlev in titlelist:
				data_tmp.append(datav[titlev['field']])
			dataformat.append(data_tmp)

		# 追加excel表数据
		for datav in dataformat:
			ws.append(datav)
		return True	

	""" 
	班级	在籍人数	参考人数	
	总分	平均分	及格率(60分以上)	优秀率(85分以上)	低分率(40分以下) 
	100及以上	99~90	89~85	84~80	79~70	69~60	
	59~50	49~40	39~30	29~20	19~10	10分以下	最低分数
	"""
	# 分段人数统计
	def fenduanPeople(self, fenshu_list, people_num):
		res = {'fen100':0,'fen90_99':0,'fen89_85':0,'fen84_80':0,'fen79_70':0,'fen69_60':0,'fen59_50':0,'fen49_40':0,'fen39_30':0,'fen29_20':0,'fen19_10':0, 'fen10xia':0,'jige':0,'youxiu':0,'difen':0, 'jigelv':0, 'youxiulv':0, 'difenlv':0}

		for i in fenshu_list:
			if i>=85:
				res['youxiu']+=1

			if i>=60:
				res['jige']+=1

			if i<=40:
				res['difen']+=1			

			if i>=100:
				res['fen100']+=1
			elif i>=90:
				res['fen90_99']+=1
			elif i>=85:
				res['fen89_85']+=1
			elif i>=80:
				res['fen84_80']+=1
			elif i>=70:
				res['fen79_70']+=1
			elif i>=60:
				res['fen69_60']+=1
			elif i>=50:
				res['fen59_50']+=1
			elif i>=40:
				res['fen49_40']+=1
			elif i>=30:
				res['fen39_30']+=1
			elif i>=20:
				res['fen29_20']+=1
			elif i>=10:
				res['fen19_10']+=1
			else:
				res['fen10xia']+=1
		
		res['jigelv'] = str(round(res['jige']/people_num*100, 2))+'%'
		res['youxiulv'] = str(round(res['youxiu']/people_num*100, 2)) + '%'
		res['difenlv'] = str(round(res['difen']/people_num*100, 2)) + '%'

		return res

	# 读取统计表格
	def readExcel(self, filename):

		wb = load_workbook(filename)
		sheet_names = wb.get_sheet_names()
		sheet = wb.get_sheet_by_name(sheet_names[0])

		rows = sheet.rows

		row_i = 0
		result = {}
		title_arr = []
		# 最大列
		max_col = 0

		# 格式化表格数据成可用格式
		for row in rows:
			line = [col.value for col in row]
			row_i+=1
			# 删除第一行 表头
			# 名字为none 没有行了
			if (line[0] is None):
				break

			# 第一行初始化表头
			if row_i == 1:
				# 循环获得最大列
				while line[max_col] is not None:
					# 第一行特殊值
					if max_col == 0:
						tmp_data = {'values':[], 'zongchengji': []}
					else:
						# 其他行，正常成绩
						tmp_data = {'values':[], 'jigenum':0, 'youxiunum':0,'difennum':0,'avgscore':0,'zongfenscore':0, 'minscore':0}
					# 标题组的值
					result[line[max_col]] = tmp_data
					title_arr.append(line[max_col])
					max_col+=1
				continue

			# 其他行成绩处理
			# 第一列是姓名
			result[title_arr[0]]['values'].append(line[0]) 

			zongchengji_tmp = 0
			for i in range(1, max_col):
				score_tmp = float(line[i]) if line[i] is not None else 0
				zongchengji_tmp += score_tmp
				result[title_arr[i]]['values'].append(score_tmp)

			result[title_arr[0]]['zongchengji'].append(zongchengji_tmp) 
			
		return (title_arr, result)

	# 格式化统计表格数据
	def formatTjTable(self, title_arr, result):
		# 数据表格
		table_list = []
		# 总人数 
		people_all = len(result[title_arr[0]]['values'])
		# 统计计算各种率
		for k,v in result.items():
			# 表头退出
			if k == title_arr[0]:
				continue

			v['kemu'] = k
			v['people'] = people_all 
			v['zongfenscore'] = sum(v['values'])
			v['avgscore'] = round(v['zongfenscore']/people_all, 2)
			v['minscore'] = min(v['values'])
			# 各分段人数
			fenduanPeople_list = self.fenduanPeople(v['values'], people_all)
			v.update(fenduanPeople_list)
			table_list.append(v)

		return table_list

def main():
	app = wx.App()
	frame = myui(None)
	frame.Show()
	app.MainLoop()

if __name__ == '__main__':
	main()